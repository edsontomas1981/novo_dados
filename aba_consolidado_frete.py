from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from utils.calcula_representacao import calcula_representacao
def cria_aba_frete_consolidado(dados, workbook,totais):
    sheet2 = workbook.create_sheet(title='Frete consolidado')

    sheet2['A1'] = 'Cnpj'
    sheet2['B1'] = 'Nome'
    sheet2['C1'] = 'Peso total'
    sheet2['D1'] = 'Valor Nfs' 
    sheet2['E1'] = 'Total Frete' 

    sheet2['F1'] = 'Percentual frete'
    sheet2['G1'] = 'Percentual Nfs' 
    sheet2['H1'] = 'Percentual Peso' 



    # Definir estilo para as células de cabeçalho
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FED766")
    # Definir cores para as linhas
    cores = ["EFF1F3", "009FB7"]  # Exemplo de cores
    num_cores = len(cores)
    linha = 2
    for cell in sheet2[1]:
        cell.font = header_font
        cell.fill = header_fill    

    iterador = 2

    for cte in dados:
        sheet2['A'+str(iterador)] = dados[cte]['dados']['tomador']
        sheet2['B'+str(iterador)] = dados[cte]['dados']['nome']
        sheet2['C'+str(iterador)] = dados[cte]['dados']['peso']
        sheet2['D'+str(iterador)] = dados[cte]['dados']['vlr_nf']
        sheet2['E'+str(iterador)] = dados[cte]['dados']['frete']
        sheet2['F'+str(iterador)] = str(round(calcula_representacao(dados[cte]['dados']['frete'],totais['frete_total']),2))+"%"
        sheet2['G'+str(iterador)] = str(round(calcula_representacao(dados[cte]['dados']['vlr_nf'],totais['vlr_nfs_total']),2))+"%"
        sheet2['H'+str(iterador)] = str(round(calcula_representacao(dados[cte]['dados']['peso'],totais['peso_total']),2))+"%"
        
        # Definir estilo para a linha
        linha_cor = cores[(linha - 2) % num_cores]
        for cell in sheet2[linha]:
            cell.fill = PatternFill(fill_type="solid", fgColor=linha_cor)

        linha += 1

        iterador += 1
