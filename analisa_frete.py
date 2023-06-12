from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

def analisa_frete(dados, workbook,totais):
    # Criar um novo arquivo Excel
    sheet1 = workbook.active
    sheet1.title = 'Analise Frete'   

    
    sheet1['A1'] = 'Tomador'
    sheet1['B1'] = 'Cte'
    sheet1['C1'] = 'Frete'
    sheet1['D1'] = 'Valor por Kg Faturado'
    sheet1['E1'] = 'Por Volume' 
    sheet1['F1'] = 'Por Vlr Nf'
    sheet1['G1'] = 'Total Peso'
    sheet1['H1'] = 'Total Frete'
    sheet1['I1'] = 'Total NF'
    sheet1['G2'] = totais['peso_total']
    sheet1['H2'] = totais['frete_total']
    sheet1['I2'] = totais['vlr_nfs_total']

    # Definir cores para as linhas
    cores = ["EFF1F3", "009FB7"]  # Exemplo de cores
    num_cores = len(cores)
    linha = 2    


    # Definir estilo para as células de cabeçalho
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FED766")
    for cell in sheet1[1]:
        cell.font = header_font
        cell.fill = header_fill   


    iterador = 2
    

    for cte in dados:
        sheet1['A'+str(iterador)] = dados[cte]['tomador']['Nome']
        sheet1['B'+str(iterador)] = cte
        sheet1['C'+str(iterador)] = round(float(dados[cte]['frete']),2)
        sheet1['D'+str(iterador)] = round(float(dados[cte]['frete'])/float(dados[cte]['peso']),2)
        sheet1['E'+str(iterador)] = round(float(dados[cte]['frete'])/float(dados[cte]['volumes']),2)
        sheet1['F'+str(iterador)] = str(round(((float(dados[cte]['frete'])/float(dados[cte]['valor'])*100)),2))+'%'
        
        iterador += 1

        # Definir estilo para a linha
        linha_cor = cores[(linha - 2) % num_cores]
        for cell in sheet1[linha]:
            cell.fill = PatternFill(fill_type="solid", fgColor=linha_cor)

        linha += 1        